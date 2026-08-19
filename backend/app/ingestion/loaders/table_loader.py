from __future__ import annotations

import csv
from io import StringIO
from pathlib import Path

from backend.app.core.models import Document, Element
from backend.app.ingestion.loaders.base import BaseLoader
from backend.app.ingestion.loaders.common import document_from_elements, normalize_text
from backend.app.ingestion.types import RouteDecision


class TableLoader(BaseLoader):
    """csv / xlsx / xls -> Document (direct, no Element pipeline required)."""

    def load(self, path: Path, route: RouteDecision, title: str | None = None) -> Document:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            text, parser = self._load_csv(path)
        elif suffix == ".xlsx":
            text, parser = self._load_xlsx(path)
        elif suffix == ".xls":
            text, parser = self._load_xls(path)
        else:
            text = path.read_text(encoding="utf-8", errors="ignore")
            parser = "table_raw"

        normalized = normalize_text(text)
        element = Element.create(
            type="table",
            text=normalized,
            table_id=path.stem,
            metadata={"parser": parser},
        )
        return document_from_elements(
            path=path,
            route=route,
            elements=[element],
            title=title or path.stem,
            extra={
                "element_type": "table",
                "table_id": path.stem,
                "actual_parser": parser,
                "loader_status": "ready",
            },
        )

    def _load_csv(self, path: Path) -> tuple[str, str]:
        raw = path.read_text(encoding="utf-8", errors="ignore")
        # Keep CSV structure; normalize via round-trip when possible.
        try:
            reader = csv.reader(StringIO(raw))
            rows = [",".join(row) for row in reader]
            return "\n".join(rows), "csv"
        except csv.Error:
            return raw, "csv_raw"

    def _load_xlsx(self, path: Path) -> tuple[str, str]:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        blocks: list[str] = []
        for sheet in workbook.worksheets:
            blocks.append(f"[sheet:{sheet.title}]")
            for row in sheet.iter_rows(values_only=True):
                cells = ["" if cell is None else str(cell) for cell in row]
                if any(cell.strip() for cell in cells):
                    blocks.append("\t".join(cells))
        workbook.close()
        return "\n".join(blocks), "openpyxl"

    def _load_xls(self, path: Path) -> tuple[str, str]:
        import xlrd

        book = xlrd.open_workbook(str(path))
        blocks: list[str] = []
        for sheet in book.sheets():
            blocks.append(f"[sheet:{sheet.name}]")
            for row_idx in range(sheet.nrows):
                cells = [str(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)]
                if any(cell.strip() for cell in cells):
                    blocks.append("\t".join(cells))
        return "\n".join(blocks), "xlrd"
